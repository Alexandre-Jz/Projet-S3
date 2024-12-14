import turtle
from math import sqrt
from random import randint
from time import time
import openpyxl
import copy 
import sys

sys.setrecursionlimit(99999)



"""Les fonctions ligne et fleche ne servent que pour afficher le graphe à l'aide de python. Elles ne sont pas importantes"""

def ligne(dist,chiffre):
    turtle.write(str(chiffre))
    turtle.dot(5)
    turtle.forward(dist)
    
def fleche(dist,chiffre,valeur):
    turtle.write(str(chiffre))
    turtle.dot(5)
    turtle.forward(dist/2)
    turtle.left(150)
    turtle.forward(20)
    turtle.back(20)
    turtle.left(60)
    turtle.forward(20)
    turtle.back(20)
    turtle.left(150)
    turtle.color('red')
    turtle.write(str(valeur))
    turtle.color('black')
    turtle.forward(dist/2)

"""La classe node est un des point d'un graphe, et la classe arbre contient donc une liste de classe node.
Les nodes sont definient par un chiffre et par la liste de ses voisins.
Cette liste prend pour arguments des sous listes avec 3 informations :
- la node voisine,
- sa distance (ATTENTION, UNIQUEMENT UTILISER POUR AFFICHER L'ARBRE != LA FONCTION d())
- sa direction, un booleen qui indique True si la liaison avec le voisin est orienté dans cette direction, False si c'est soit non orienté, soit dans l'autre direction"""

class node():
    def __init__(self, chiffre):
        self.chiffre=chiffre
        self.lvoisins=[]
    


    def add_voisins(self, node, distance, direction): 
        self.lvoisins.append([node, distance, False])
        node.lvoisins.append([self, distance, direction])

"""La classe Arbre est la plus importante, Les commentaires seront donc sur les lignes en question où il peut etre compliquer de comprendre"""

class Arbre():
    def __init__(self, lnodes=[], lchemins=[], dictvaleurschemins={}):
        self.lnodes=lnodes #liste des nodes
        self.lchemins=lchemins #liste des chemins (uniquement les chemins qui existent, pas orienté)
        self.dictvaleurschemins=dictvaleurschemins #liste des valeurs des chemins (meme de ceux qui n'existe pas, utilisé pour la fonction d)



    def add_node(self, chiffre_node, chiffre_voisin=None, distance=100, direction=False): #ajoute une node à l'arbre en rafrachisant les informations de la node voisine

        """ajoute la node"""

        a=node(chiffre_node)
        self.lnodes.append(a)
        if chiffre_voisin!=None:
            for nod in self.lnodes:
                if nod.chiffre==chiffre_voisin:
                    a.add_voisins(nod, distance, direction)


        if chiffre_voisin!=None:

            """ajoute le chemin"""

            self.lchemins.append([chiffre_node, chiffre_voisin])

            """ajoute les valeurs des chemins associé à la node rajouté, pour le moment aléatoire entre 1 et 100"""

            for n in range(len(self.lnodes)-1): 
                nod=self.lnodes[n].chiffre
                valeur=randint(1,99) #modification possible en fonction de la dissimilarité recherchée
                self.dictvaleurschemins[chiffre_node,nod]=valeur
                self.dictvaleurschemins[nod,chiffre_node]=valeur



    def add_fleche(self, node1_chiffre, node2_chiffre): #ajoute une direction entre deux nodes

        for node in self.lnodes:
            if node.chiffre==node1_chiffre:
                for voisins in node.lvoisins:
                    if voisins[0].chiffre==node2_chiffre:
                        voisins[2]=True
                        


    def add_random_fleche(self, seed=False): #prend un arbre sans direction et rajoute sur chaque chemin une direction aléatoire ou ajoute en fonction de la seed

        if seed==False:

            """ajout aléatoire"""
            
            for chemin in self.lchemins:
                if randint(0,1)==0:
                    self.add_fleche(chemin[0], chemin[1])
                else:
                    self.add_fleche(chemin[1], chemin[0])

        else:

            """ajout controlé par la seed. la seed est une liste de longueur egale au nombre de chemin qui prend pour valeur :
            - 1 si le chemin doit etre orienté dans le sens normal
            - 0 pour le sens opposé"""

            for a in range(len(seed)):
                val=seed[a]
                chemin=self.lchemins[a]
                if val=='1':
                    self.add_fleche(chemin[0], chemin[1])
                else:
                    self.add_fleche(chemin[1], chemin[0])



    def d(self,chiffre1,chiffre2): #la fonction d qui represente la dismilarité rend simplement cette valeur entre deux nodes
        
        return(self.dictvaleurschemins[chiffre1,chiffre2])
    


    def compte_resolution(self): #prend un arbre orienté en entrée et rend son nombre de chemin ou None si un chemin n'est pas valide

        lchemintot=[]

        """Premierement on prend les chemins de taille 1 que l'on met dans la liste lchemintot avec la bonne orientation (node1,node2!=node2,node1)"""

        for chemin in self.lchemins:
            for node in self.lnodes:
                if node.chiffre==chemin[0]:
                    for voisin in node.lvoisins:
                        if voisin[0].chiffre==chemin[1]:
                            if voisin[2]==True:
                                lchemintot.append((node.chiffre,voisin[0].chiffre))
                            elif voisin[2]==False:
                                lchemintot.append((voisin[0].chiffre,node.chiffre))
        
        """Mantenant on parcoure la liste et lorsque un chemin ce termine à l'endroit ou un autre debute,
        on rajoute le chemin englobant les deux chemins à la liste de chemin et on teste si elle respecte la fonction.
        On recommence ça autant de fois que necessaire jusqu'a avoir tout les chemins, ou avoir la fonction non respecté."""

        a=len(lchemintot)
        for i in range(a):
            lchemintot=list(set(lchemintot))
            for chem in lchemintot:
                for chem2 in lchemintot:
                    if chem[1]==chem2[0]:
                        lchemintot.append((chem[0],chem2[1]))
                        if self.d(chem[0],chem2[1])<self.d(chem[0],chem[1]) or self.d(chem[0],chem2[1])<self.d(chem2[0],chem2[1]):
                            return(None)


        """Pour eviter les duplication, on transforme la liste en set avant de rendre sa taille = le nombre de chemin de l'arbre"""

        return len(set(lchemintot))


    
    def resolution_brute_force(self): #La fonction principale du programme, le brute force.
        max=0
        for a in range(2**len(self.lchemins)):

            """cree toute les seeds possible (utilisation des nombres binaires car ils parcourent tout de 00000000 à 11111111)"""

            seed=bin(a).replace("0b", "")
            seed=list(seed)
            while len(seed)!=len(self.lchemins):
                seed.insert(0,'0')

            """Pour chaque seed, crée une copie de l'arbre principale, l'oriente en fonction de la seed puis compte le nombre de chemin"""
            arbre_copy=copy.deepcopy(self)
            arbre_copy.add_random_fleche(seed)
            max_temp=arbre_copy.compte_resolution()

            """rend le maximum et l'arbre associé"""

            if max_temp != None and max_temp>max:
                max=max_temp
                arbre_maximal=copy.deepcopy(arbre_copy)
            
        return(max,arbre_maximal)


    def excel(self):

        workbook = openpyxl.Workbook()

        sheet = workbook.active
        sheet.title = "Taille " + str(len(self.lnodes))

        header = ["ID", "Seed", "Nombre de chemin trouve", "Temps de calcul", "Arbre : " + str(self.lchemins), "Temps Total"]

        for col_num, column_title in enumerate(header, 1):
            sheet.cell(row=1, column=col_num, value=column_title)

        current_row=2
        lenprevseed=0

        depart_tot=time()
        for a in range(2**len(self.lchemins)):

            """cree toute les seeds possible (utilisation des nombres binaires car ils parcourent tout de 00000000 à 11111111)"""

            seed=bin(a).replace("0b", "")
            lseed=list(seed)
            while len(lseed)!=len(self.lchemins):
                lseed.insert(0,'0')

            """Pour chaque seed, crée une copie de l'arbre principale, l'oriente en fonction de la seed puis compte le nombre de chemin"""

            depart=time()

            arbre_copy=copy.deepcopy(self)
            arbre_copy.add_random_fleche(seed)
            max_temp=arbre_copy.compte_resolution()

            durée=time()-depart

            """rend le tableau excel"""

            row_data=[a+1,seed,max_temp,durée]

            for col_num, cell_value in enumerate(row_data, 1):
                sheet.cell(row=current_row, column=col_num, value=cell_value)
            current_row += 1
            if len(seed)!=lenprevseed:
                print(seed)
            lenprevseed=len(seed)
        print(self.lnodes)

        sheet.cell(row=2, column=6, value=time()-depart_tot)

        # Sauvegarde du fichier Excel
        file_name = "donnees_arbre"+str(len(self.lnodes))+".xlsx"
        workbook.save(file_name)
        print(f"Fichier Excel sauvegardé sous le nom : {file_name}")
        return


    def affiche(self): #Fonction utile que pour afficher les arbres, compliqué et pas importante pour le probleme principal
        turtle.hideturtle()
        turtle.speed(0)
        arbre_copy=copy.deepcopy(self)
        for nod in arbre_copy.lnodes:
            nod.angle=360/len(nod.lvoisins)
        nod=arbre_copy.lnodes[0]
        while nod.lvoisins!=[]:
            turtle.left(nod.angle)
            if nod.lvoisins[-1][2]==True:
                fleche(nod.lvoisins[-1][1],nod.chiffre,self.d(nod.chiffre,nod.lvoisins[-1][0].chiffre))
            else:
                ligne(nod.lvoisins[-1][1],nod.chiffre)
            new_nod=nod.lvoisins[-1][0]
            nod.lvoisins.pop(-1)
            nod=new_nod
            turtle.left(180)
        turtle.done()

def create_rand_tree(n):
    arb=Arbre([],[],{})
    arb.add_node(1)
    arb.add_node(2,1)
    for i in range(3,n+1):
        arb.add_node(i,randint(1,i-1))
    return arb

def excel_debut_fin(a,b):
    for i in range(a,b+1):
        arb=create_rand_tree(i)
        arb.excel()

a=create_rand_tree(8)
a.add_random_fleche()
a.affiche()
